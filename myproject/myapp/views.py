
from django.shortcuts import render
from decimal import Decimal

# Create your views here.
from django.shortcuts import render, get_object_or_404, redirect
from django.http import HttpResponse, HttpResponseRedirect
from .models import User, Bus, Book, Seat
from django.contrib.auth import authenticate, login, logout, get_user_model
from django.contrib.auth.models import User
from django.contrib.auth.decorators import user_passes_test, login_required
from django.core.exceptions import PermissionDenied
from .forms import UserLoginForm, UserRegisterForm
from decimal import Decimal
from datetime import datetime
from django.utils.crypto import get_random_string
from django.utils.timezone import now
from django.template.loader import get_template
import os
from django.conf import settings
from xhtml2pdf import pisa
from io import BytesIO
from django.http import JsonResponse
from myapp.models import Bus, Book, Seat, Blog, TravelGallery, BlogForm, BusForm, TravelGalleryForm, Review, ReviewForm #later add forms in another file
import string
import csv
import io
import xlsxwriter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
import openpyxl
from openpyxl.utils import get_column_letter
from django.contrib import messages
from django.db import transaction
# Assume seat_labels already created like ['A1', 'A2', ..., 'J4']
from itertools import zip_longest

def is_admin(user):
    if not user.is_authenticated or not (user.is_staff or user.is_superuser):
        raise PermissionDenied
    return True

def home(request):
    if request.user.is_authenticated:
        reviews = Review.objects.select_related('bus').order_by('-created_at')[:20]  # Fetch latest 20 reviews
        return render(request, 'myapp/Nav Bar/home.html', {'reviews': reviews})
    else:
        return render(request, 'myapp/Nav Bar/signin.html')


def about_us(request):
    return render(request, 'myapp/Quick links/about_us.html')


@login_required(login_url='signin')
def findbus(request):
    context = {}

    if request.method == 'POST':
        # Normalize input: strip whitespace and lowercase
        source_r = request.POST.get('source', '').strip().lower()
        dest_r = request.POST.get('destination', '').strip().lower()
        date_r = request.POST.get('date')
        selected_bus_id = request.POST.get('selected_bus_id')

        # Use __iexact for case-insensitive matching
        bus_list = Bus.objects.filter(
            source__iexact=source_r,
            dest__iexact=dest_r,
            date=date_r
        )

        if bus_list:
            if selected_bus_id:
                selected_bus = Bus.objects.get(id=selected_bus_id)
            else:
                selected_bus = bus_list.first()

            booked_seats = []
            seat_labels = []

            if selected_bus:
                # Get total number of seats
                total_seats = int(selected_bus.nos)

                # Generate seat labels like A1, A2, ..., B1, B2, ...
                seat_labels = []
                rows = string.ascii_uppercase  # 'A', 'B', ...
                seat_num = 0
                for r in rows:
                    for c in ['1', '2', '3', '4']:  # 4 seats per row
                        seat_num += 1
                        if seat_num > total_seats:
                            break
                        seat_labels.append(r + c)
                    if seat_num >= total_seats:
                        break

                # Get booked seats
                booked = Seat.objects.filter(bus=selected_bus, date=date_r, is_booked=True)
                booked_seats = [seat.seat_number for seat in booked]

            context['bus_list'] = bus_list
            context['selected_bus_id'] = selected_bus.id
            context['selected_date'] = date_r
            context['booked_seats'] = booked_seats
            context['seat_labels'] = seat_labels
            # Group seats into rows of 4 for 2+aisle+2 layout
            seat_rows = [seat_labels[i:i + 4] for i in range(0, len(seat_labels), 4)]

            context['seat_rows'] = seat_rows
            context['seat_labels'] = seat_labels  # Pass to template

            return render(request, 'myapp/list.html', context)
        else:
            context["error"] = "Sorry no buses available"
            return render(request, 'myapp/Nav Bar/findbus.html', context)

    else:
        return render(request, 'myapp/Nav Bar/findbus.html')


@login_required(login_url='signin')
def get_booked_seats(request):
    bus_id = request.GET.get('bus_id')
    date = request.GET.get('date')

    if bus_id and date:
        try:
            bus = Bus.objects.get(id=bus_id)
            total_seats = int(bus.nos)

            seat_labels = []
            rows = string.ascii_uppercase
            seat_num = 0
            for r in rows:
                for c in ['1', '2', '3', '4']:
                    seat_num += 1
                    if seat_num > total_seats:
                        break
                    seat_labels.append(r + c)
                if seat_num >= total_seats:
                    break

            booked = Seat.objects.filter(bus=bus, date=date, is_booked=True)
            booked_seats = [seat.seat_number for seat in booked]

            return JsonResponse({
                'seat_labels': seat_labels,
                'booked_seats': booked_seats,
            })
        except Bus.DoesNotExist:
            return JsonResponse({'error': 'Bus not found'}, status=404)
    return JsonResponse({'error': 'Invalid request'}, status=400)



@login_required(login_url='signin')
def bookings(request):
    context = {}
    if request.method == 'POST':
        bus_id = request.POST.get('bus_id')
        selected_seats = request.POST.getlist('seats')
        user = request.user

        if not selected_seats:
            context["error"] = "Please select at least one seat"
            return render(request, 'myapp/Nav Bar/findbus.html', context)

        try:
            bus = Bus.objects.get(id=bus_id)
        except Bus.DoesNotExist:
            context["error"] = "Invalid bus ID"
            return render(request, 'myapp/Nav Bar/findbus.html', context)

        already_booked = Seat.objects.filter(bus=bus, seat_number__in=selected_seats, date=bus.date, is_booked=True)
        if already_booked.exists():
            context["error"] = "Some selected seats are already booked. Please refresh and try again."
            return render(request, 'myapp/Nav Bar/findbus.html', context)

        book = Book.objects.create(
            name=user.username,
            email=user.email,
            userid=user.id,
            busid=bus.id,
            bus_name=bus.bus_name,
            source=bus.source,
            dest=bus.dest,
            nos=len(selected_seats),
            price=bus.price,
            date=bus.date,
            time=bus.time,
            total_price=len(selected_seats) * bus.price,
            status='BOOKED'
        )

        for seat in selected_seats:
            seat_obj, created = Seat.objects.get_or_create(
                bus=bus,
                seat_number=seat,
                date=bus.date,
                defaults={
                    'is_booked': True,
                    'booked_by': user,
                    'booking': book
                }
            )

            if not created:
                # If seat already exists, update it for the new booking
                seat_obj.is_booked = True
                seat_obj.booked_by = user
                seat_obj.booking = book
                seat_obj.save()


        bus.rem -= len(selected_seats)
        bus.save()

        # Calculate total cost
        cost = book.price * book.nos  # price per seat * number of seats

        return render(request, 'myapp/bookings.html', {'book': book, 'seats': selected_seats, 'cost': cost})

    return render(request, 'myapp/Nav Bar/findbus.html')


@login_required(login_url='signin')
def cancellings(request):
    context = {}
    if request.method == 'POST':
        id_r = request.POST.get('bus_id')  # This is actually the booking ID (not bus ID!)

        try:
            book = Book.objects.get(id=id_r)
            bus = Bus.objects.get(id=book.busid)

            # Free up the seats: set is_booked=False for associated seats
            Seat.objects.filter(booking=book).update(is_booked=False)

            # Or delete the seat entries if you don't want to keep them
            # Seat.objects.filter(booking=book).delete()

            # Update bus remaining seats
            bus.rem += book.nos
            bus.save()

            # Cancel the booking
            book.status = 'CANCELLED'
            book.total_price = 0
            book.save()

            return redirect(seebookings)

        except Book.DoesNotExist:
            context["error"] = "Sorry, you have not booked that bus"
            return render(request, 'myapp/error.html', context)
    else:
        return render(request, 'myapp/Nav Bar/findbus.html')



@login_required(login_url='signin')
def seebookings(request):
    id_r = request.user.id
    book_list = Book.objects.filter(userid=id_r)
    
    if book_list:
        # Add seat numbers to each booking
        booking_data = []
        for booking in book_list:
            seats = Seat.objects.filter(booking=booking).values_list('seat_number', flat=True)
            # total_cost = float(booking.price) * len(list(seats))
            booking_data.append({
                'booking': booking,
                'seat_numbers': list(seats),
                'total_cost': booking.total_price,
            })
        return render(request, 'myapp/Nav Bar/booklist.html', {'booking_data': booking_data})
    else:
        context = {"error": "Sorry no buses booked"}
        return render(request, 'myapp/Nav Bar/findbus.html', context)
    

@login_required(login_url='signin')
def get_suggestions(request):
    query = request.GET.get('term', '').lower()
    exclude = request.GET.get('exclude', '').lower()

    locations = [
        "Dhaka", "Chittagong", "Rajshahi", "Khulna", "Barisal",
        "Sylhet", "Rangpur", "Mymensingh", "Coxsbazar", "Saint Martin", "Jessore", "Noakhali",
    ]

    # Filter by query and exclude
    filtered = [loc for loc in locations if query in loc.lower() and loc.lower() != exclude]
    return JsonResponse(filtered[:10], safe=False)


def invoice_verify_view(request):
    if request.method == 'POST':
        booking_id = request.POST.get('booking_id')
        status = 'INVALID'
        try:
            booking = Book.objects.get(id=booking_id)
            status = booking.status
        except Book.DoesNotExist:
            pass

        # Store status temporarily in session and redirect
        request.session['status'] = status
        return redirect("verify_invoice")  # or your URL name

    # Handle GET: show status if redirected after POST
    status = request.session.pop('status', None)
    context = {
        'checked': status is not None,
        'status': status
    }
    return render(request, 'myapp/Quick links/invoice_verify.html', context)


@user_passes_test(is_admin)
def manage_gallery(request):
    galleries = TravelGallery.objects.all().order_by('-date_added')
    return render(request, 'myapp/Manage/manage_gallery.html', {'galleries': galleries})


@user_passes_test(is_admin)
def create_or_edit_gallery(request, pk=None):
    gallery = get_object_or_404(TravelGallery, pk=pk) if pk else None
    if request.method == 'POST':
        form = TravelGalleryForm(request.POST, request.FILES, instance=gallery)
        if form.is_valid():
            form.save()
            return redirect('manage_gallery')
    else:
        form = TravelGalleryForm(instance=gallery)
    return render(request, 'myapp/Manage/create_or_edit_gallery.html', {'form': form, 'gallery': gallery})


@user_passes_test(is_admin)
def delete_gallery(request, pk):
    gallery = get_object_or_404(TravelGallery, pk=pk)

    # Delete the associated image file from media storage
    if gallery.photo:
        image_path = os.path.join(settings.MEDIA_ROOT, gallery.photo.name)
        if os.path.isfile(image_path):
            os.remove(image_path)

    gallery.delete()
    return redirect('manage_gallery')


@user_passes_test(is_admin)
def manage_blogs(request):
    blogs = Blog.objects.all().order_by('-date_posted')
    return render(request, 'myapp/Manage/manage_blogs.html', {'blogs': blogs})


@user_passes_test(is_admin)
def edit_blog(request, blog_id=None):
    if blog_id:
        blog = get_object_or_404(Blog, id=blog_id)
    else:
        blog = None

    if request.method == 'POST':
        form = BlogForm(request.POST, instance=blog)
        if form.is_valid():
            form.save()
            return redirect('manage_blogs')
    else:
        form = BlogForm(instance=blog)

    return render(request, 'myapp/Manage/edit_blog.html', {
        'form': form,
        'preview': form.instance if blog else None,
    })


@user_passes_test(is_admin)
def delete_blog(request, blog_id):
    blog = get_object_or_404(Blog, id=blog_id)

    if request.method == "POST":
        blog.delete()
        return redirect('manage_blogs')
    
    # Deny GET requests for security
    raise PermissionDenied


@user_passes_test(is_admin)
def admin_bus_list(request):
    buses = Bus.objects.all().order_by('-date')
    return render(request, 'myapp/Manage/manage_buses.html', {'buses': buses})


@user_passes_test(is_admin)
def add_bus(request):
    if request.method == 'POST':
        form = BusForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('admin_bus_list')
    else:
        form = BusForm()
    return render(request, 'myapp/Manage/add_or_edit_bus.html', {'form': form, 'title': 'Add New Bus'})


@user_passes_test(is_admin)
def edit_bus(request, bus_id):
    bus = get_object_or_404(Bus, id=bus_id)
    if request.method == 'POST':
        form = BusForm(request.POST, instance=bus)
        if form.is_valid():
            form.save()
            return redirect('admin_bus_list')
    else:
        form = BusForm(instance=bus)
    return render(request, 'myapp/Manage/add_or_edit_bus.html', {'form': form, 'title': f'Edit {bus.bus_name}'})


@user_passes_test(is_admin)
def delete_bus(request, bus_id):
    if request.method == 'POST':
        bus = get_object_or_404(Bus, id=bus_id)
        bus.delete()
        messages.success(request, 'Bus deleted successfully.')
    else:
        messages.error(request, 'Invalid request method.')
    return redirect('admin_bus_list')


def generate_all_seats_for_bus_date(bus, date):
    total_seats = int(bus.nos)
    seat_labels = []
    rows = string.ascii_uppercase
    seat_num = 0
    for r in rows:
        for c in ['1', '2', '3', '4']:
            seat_num += 1
            if seat_num > total_seats:
                break
            seat_label = r + c
            seat_labels.append(seat_label)
        if seat_num >= total_seats:
            break

    for seat_label in seat_labels:
        Seat.objects.get_or_create(
            bus=bus,
            date=date,
            seat_number=seat_label,
            defaults={'is_booked': False}
        )


@user_passes_test(is_admin)
def admin_manage_bookings(request):
    bookings = Book.objects.all().order_by('-date', '-time')
    return render(request, 'myapp/Manage/manage_bookings.html', {'bookings': bookings})


@user_passes_test(is_admin)
def delete_booking(request, booking_id):
    booking = get_object_or_404(Book, id=booking_id)

    if booking.status != "CANCELLED":
        messages.error(request, 'Only cancelled bookings can be deleted.')
        return redirect('admin_manage_bookings')

    # Delete associated seat records
    Seat.objects.filter(booking=booking).delete()

    # Optionally update bus.rem count — though unnecessary if booking is already cancelled
    # We skip this assuming rem was adjusted during cancel/rebook

    booking.delete()
    messages.success(request, 'Booking and related seat records successfully deleted.')
    return redirect('admin_manage_bookings')


@user_passes_test(is_admin)
def cancel_booking(request, booking_id):
    print(f"Cancel booking view called with booking_id={booking_id}")  # DEBUG LINE
    booking = get_object_or_404(Book, id=booking_id)

    if booking.status == "BOOKED":
        print(f"Current booking status: {booking.status}")  # DEBUG
        try:
            with transaction.atomic():
                booking.total_price = 0
                booking.status = "CANCELLED"
                booking.save()

                seats = Seat.objects.filter(booking=booking)
                seat_count = seats.count()
                print(f"Seats found for booking: {seats.count()}")  # DEBUG

                # Update seats as unbooked
                seats.update(is_booked=False, booking=None, booked_by=None)

                # Get bus object (assuming booking.busid is an integer FK)
                bus = Bus.objects.get(id=booking.busid)

                # Update remaining seats count on bus
                bus.rem = bus.rem + seat_count
                bus.save()

            messages.success(request, 'Booking cancelled successfully.')
        except Exception as e:
            messages.error(request, f'Error cancelling booking: {str(e)}')
    else:
        messages.warning(request, 'Booking is already cancelled or not active.')

    return redirect('admin_manage_bookings')



@user_passes_test(is_admin)
def rebook_booking(request, booking_id):
    User = get_user_model()
    booking = get_object_or_404(Book, id=booking_id)

    if booking.status != "CANCELLED":
        context = {'error_message': "Only cancelled bookings can be rebooked."}
        messages.error(request, 'Only cancelled bookings can be rebooked.')
        return render(request, 'myapp/error.html', context)

    bus = get_object_or_404(Bus, id=booking.busid)
    date = booking.date
    seats_to_book = int(booking.nos)

    with transaction.atomic():
        generate_all_seats_for_bus_date(bus, date)

        available_seats = Seat.objects.select_for_update().filter(
            bus=bus,
            date=date,
            is_booked=False
        )[:seats_to_book]

        if available_seats.count() < seats_to_book:
            context = {'error_message': "Not enough seats available to rebook."}
            messages.error(request, 'Not enough seats available to rebook.')
            return render(request, 'myapp/error.html', context)

        try:
            passenger_user = User.objects.get(email=booking.email)
        except User.DoesNotExist:
            passenger_user = None

        booked_seat_numbers = []
        for seat in available_seats:
            seat.is_booked = True
            seat.booking = booking
            seat.booked_by = passenger_user
            seat.save()
            booked_seat_numbers.append(seat.seat_number)

        # Update booking info
        booking.status = "BOOKED"
        booking.nos = len(booked_seat_numbers)
        booking.total_price = booking.nos * booking.price

        if hasattr(booking, 'seat_numbers'):
            booking.seat_numbers = ",".join(booked_seat_numbers)

        booking.save()

        # Update bus remaining seats count
        bus.rem -= booking.nos
        bus.save()

        context = {'error_message': "Booking Successfully rebooked with new available seats."}
        messages.success(request, 'Booking successfully rebooked with new available seats.')

    return render(request, 'myapp/error.html', context)





@user_passes_test(is_admin)
def create_booking(request):
    User = get_user_model()  # <- get User model here once at the start

    if request.method == 'POST':
        name = request.POST['name']
        email = request.POST['email']
        busid = int(request.POST['busid'])
        nos = int(request.POST['nos'])
        seat_input = request.POST.get('seat_numbers', '').strip()

        bus = get_object_or_404(Bus, id=busid)
        bus_name = bus.bus_name
        source = bus.source
        dest = bus.dest
        date = bus.date
        time = bus.time
        price = bus.price

        # Validate and assign seat numbers
        with transaction.atomic():
            generate_all_seats_for_bus_date(bus, date)
            if seat_input:
                seat_numbers = [s.strip().upper() for s in seat_input.split(',') if s.strip()]
                if len(seat_numbers) != nos:
                    return HttpResponse("Number of seats doesn't match seat numbers provided.")
                available = Seat.objects.select_for_update().filter(
                    bus=bus, date=date, seat_number__in=seat_numbers, is_booked=False
                )
                if available.count() != nos:
                    return HttpResponse("Some requested seats are already booked or invalid.")
            else:
                available = Seat.objects.select_for_update().filter(
                    bus=bus, date=date, is_booked=False
                )[:nos]
                if available.count() < nos:
                    return HttpResponse("Not enough seats available.")
                seat_numbers = [s.seat_number for s in available]

            passenger_user = None
            try:
                passenger_user = User.objects.get(email=email)
            except User.DoesNotExist:
                passenger_user = None

            random_password = get_random_string(length=8)
            if passenger_user is None:
                passenger_user = User.objects.create_user(
                    email=email,
                    username=email,
                    password=random_password
                )

            # Create booking
            booking = Book.objects.create(
                name=name,
                email=email,
                userid=passenger_user.id if passenger_user else 0,
                busid=bus.id,
                bus_name=bus.bus_name,
                source=bus.source,
                dest=bus.dest,
                nos=nos,
                price=bus.price,
                date=bus.date,
                time=bus.time,
                total_price=nos * bus.price,
                status='BOOKED'
            )

            # Mark seats as booked
            for seat_number in seat_numbers:
                seat = Seat.objects.get(bus=bus, date=date, seat_number=seat_number)
                seat.is_booked = True
                seat.booking = booking
                seat.booked_by = passenger_user
                seat.save()

            # Update the remaining seats in the bus
            bus.rem -= nos
            bus.save()

            return redirect('admin_manage_bookings')

    return HttpResponse("Invalid request")


def export_bookings_csv(bookings):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="bookings.csv"'

    writer = csv.writer(response)
    writer.writerow(['Booking ID', 'User Name', 'Bus ID', 'Bus Name', 'Route From', 'Route To', 'Seats', 'Price', 'Total', 'Date'])

    for b in bookings:
        writer.writerow([b.id, b.name, b.busid, b.bus_name, b.source, b.dest, b.nos, f'{b.price:.2f}', f'{b.nos * b.price:.2f}', b.date])

    return response


def export_bookings_excel(bookings):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bookings"

    headers = ['Booking ID', 'STATUS', 'User Name', 'Bus ID', 'Bus Name', 'Route From', 'Route To', 'Seats', 'Price', 'Total', 'Date']
    ws.append(headers)

    for b in bookings:
        ws.append([b.id, b.status, b.name, b.busid, b.bus_name, b.source, b.dest, b.nos, float(f'{b.price:.2f}'), float(f'{b.nos * b.price:.2f}'), str(b.date)])

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="bookings.xlsx"'
    wb.save(response)
    return response


def export_bookings_pdf(bookings):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="bookings.pdf"'

    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter

    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(1 * inch, height - 1 * inch, "Bookings Report")

    # Table header
    p.setFont("Helvetica-Bold", 10)
    headers = ['Booking ID', 'STATUS', 'User Name', 'Bus ID', 'Bus Name', 'Route From', 'Route To', 'Seats', 'Price', 'Total', 'Date']
    x_offset = 50
    y_offset = height - 1.5 * inch
    col_widths = [60, 60, 80, 50, 80, 70, 70, 40, 40, 50, 60]

    for i, header in enumerate(headers):
        p.drawString(x_offset, y_offset, header)
        x_offset += col_widths[i]
    
    # Table rows
    p.setFont("Helvetica", 9)
    y_offset -= 15
    for b in bookings:
        x_offset = 50
        row = [
            str(b.id), b.status, b.name, str(b.busid), b.bus_name, b.source, b.dest,
            str(b.nos), f'{b.price:.2f}', f'{b.nos * b.price:.2f}', str(b.date)
        ]
        for i, cell in enumerate(row):
            p.drawString(x_offset, y_offset, cell)
            x_offset += col_widths[i]
        y_offset -= 15
        if y_offset < 50:  # create new page if running out of space
            p.showPage()
            y_offset = height - 50

    p.showPage()
    p.save()
    return response


@user_passes_test(is_admin)
@user_passes_test(is_admin)
def admin_booking_view(request):
    query = request.GET
    bookings = Book.objects.all()  # show all by default
    total_amount = 0

    date = query.get('date')
    if date:
        bookings = bookings.filter(date=date)

    filter_by = query.get('filter_by')
    if filter_by == 'booking_id':
        start = query.get('booking_id_start')
        end = query.get('booking_id_end')
        if start and end:
            bookings = bookings.filter(id__range=(start, end))
        elif start:
            bookings = bookings.filter(id__gte=start)
        elif end:
            bookings = bookings.filter(id__lte=end)

    elif filter_by == 'bus_id':
        bookings = bookings.filter(busid=query.get('bus_id'))

    elif filter_by == 'bus_name':
        bookings = bookings.filter(bus_name__icontains=query.get('bus_name'))

    elif filter_by == 'route':
        source = query.get('source')
        destination = query.get('destination')
        if source:
            bookings = bookings.filter(source__icontains=source)
        if destination:
            bookings = bookings.filter(dest__icontains=destination)

    elif filter_by == 'username':
        bookings = bookings.filter(name__icontains=query.get('username'))

    # Calculate total only for visible records
    total_amount = sum([booking.total_price for booking in bookings])

    # Export (only what is currently shown — filtered or all)
    export_format = query.get('export')
    if export_format == 'csv':
        return export_bookings_csv(bookings)
    elif export_format == 'excel':
        return export_bookings_excel(bookings)
    elif export_format == 'pdf':
        return export_bookings_pdf(bookings)

    # Render page normally
    context = {
        'bookings': bookings,
        'total_amount': total_amount,
    }
    return render(request, 'myapp/Manage/manage_booking_records.html', context)


@login_required(login_url='signin')
def bus_detail(request, bus_id):
    bus = get_object_or_404(Bus, id=bus_id)
    reviews = Review.objects.filter(bus=bus).order_by('-created_at')  # show reviews too
    return render(request, 'myapp/bus_detail.html', {'bus': bus, 'reviews': reviews})


@login_required
def add_review(request, bus_id):
    bus = get_object_or_404(Bus, id=bus_id)

    if request.method == 'POST':
        form = ReviewForm(request.POST)
        if form.is_valid():
            review = form.save(commit=False)
            review.bus = bus
            # auto-fill from logged-in user
            review.user = request.user
            review.name = request.user.get_full_name() or request.user.username
            review.email = request.user.email
            review.save()
            return redirect('home')  # or where you want
    else:
        # Pre-fill form fields if user is logged in
        initial_data = {
            'name': request.user.get_full_name() or request.user.username,
            'email': request.user.email
        }
        form = ReviewForm(initial=initial_data)

    return render(request, 'myapp/add_review.html', {'form': form, 'bus': bus})

def clean_old_buses():
    old_buses = Bus.objects.filter(date__lt=now().date())
    for bus in old_buses:
        if not Book.objects.filter(busid=bus.id).exists():
            bus.delete()

def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


@login_required(login_url='signin')
def download_invoice(request, booking_id):
    book = Book.objects.get(id=booking_id)
    cost = book.nos * book.price  # calculate total cost

    # Get all seats associated with this booking
    seats = Seat.objects.filter(booking=book).values_list('seat_number', flat=True)

    context = {
        'book': book,
        'cost': cost,
        'now': now(),
        'seats': list(seats),  # Pass the list of seat numbers
    }

    pdf = render_to_pdf('myapp/invoice.html', context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f"Invoice_{book.id}.pdf"
        content = f"inline; filename={filename}"
        response['Content-Disposition'] = content
        return response
    return HttpResponse("Error generating PDF", status=500)


def blog_list(request):
    blogs = Blog.objects.order_by('-date_posted')
    return render(request, 'myapp/Quick links/blog.html', {'blogs': blogs})

def gallery_list(request):
    photos = TravelGallery.objects.all().order_by('-id')
    return render(request, 'myapp/Quick links/travel_gallery.html', {'photos': photos})

def gallery_detail(request, pk):
    photo = get_object_or_404(TravelGallery, pk=pk)
    return render(request, 'myapp/gallery_detail.html', {'photo': photo})

def cookies_view(request):
    return render(request, 'myapp/Support/cookies.html')

def legal_notice_view(request):
    return render(request, 'myapp/Support/legal_notice.html')

def privacy_policy_view(request):
    return render(request, 'myapp/Support/privacy_policy.html')

def terms(request):
    return render(request, 'myapp/Support/terms.html')

def refund(request):
    return render(request, 'myapp/Support/refund.html')


def signup(request):
    context = {}
    if request.method == 'POST':
        name_r = request.POST.get('name')
        email_r = request.POST.get('email')
        password_r = request.POST.get('password')
        user = User.objects.create_user(name_r, email_r, password_r, )
        if user:
            login(request, user)
            return render(request, 'myapp/thank.html')
        else:
            context["error"] = "Provide valid credentials"
            return render(request, 'myapp/Nav Bar/signup.html', context)
    else:
        return render(request, 'myapp/Nav Bar/signup.html', context)


def signin(request):
    context = {}
    if request.method == 'POST':
        name_r = request.POST.get('name')
        password_r = request.POST.get('password')
        user = authenticate(request, username=name_r, password=password_r)
        if user:
            login(request, user)
            # username = request.session['username']
            context["user"] = name_r
            context["id"] = request.user.id
            return render(request, 'myapp/success.html', context)
            # return HttpResponseRedirect('success')
        else:
            context["error"] = "Provide valid credentials"
            return render(request, 'myapp/Nav Bar/signin.html', context)
    else:
        context["error"] = "You are not logged in"
        return render(request, 'myapp/Nav Bar/signin.html', context)


def signout(request):
    context = {}
    logout(request)
    context['error'] = "You have been logged out"
    return render(request, 'myapp/Nav Bar/signin.html', context)


def success(request):
    context = {}
    context['user'] = request.user
    return render(request, 'myapp/success.html', context)
